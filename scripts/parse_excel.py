import openpyxl
import json

wb = openpyxl.load_workbook("docs/parametro iniciales/maestro de paradas.xlsx")
sheet = wb.active

rows = []
# Headers are: CATEGORIA, TIEMPO MUERTO, MOTIVO, LAYOUT SECADERO
# Column 1: CATEGORIA
# Column 2: TIEMPO MUERTO
# Column 3: MOTIVO
# Column 4: LAYOUT SECADERO
for r in range(2, sheet.max_row + 1):
    cat = sheet.cell(row=r, column=1).value
    tm = sheet.cell(row=r, column=2).value
    motivo = sheet.cell(row=r, column=3).value
    layout = sheet.cell(row=r, column=4).value
    if cat or tm or motivo or layout:
        rows.append({
            "categoria": str(cat).strip() if cat else "",
            "tiempo_muerto": str(tm).strip() if tm else "",
            "motivo": str(motivo).strip() if motivo else "",
            "layout": str(layout).strip() if layout else ""
        })

with open("database/maestro_paradas.json", "w", encoding="utf-8") as f:
    json.dump(rows, f, ensure_ascii=False, indent=2)

print("Parsed", len(rows), "rows and saved to database/maestro_paradas.json")
